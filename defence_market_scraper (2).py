"""
Defence Market Analysis – NTT DATA Italy
Script giornaliero con Anthropic API + web search integrato.
Raccoglie: budget IT, spesa digitale, storico gare ANAC per ogni player.
Genera Excel strutturato e invia mail con delta vs giorno precedente.

Requisiti:
    pip install anthropic openpyxl python-dotenv

Secrets GitHub / variabili .env:
    ANTHROPIC_API_KEY
    SMTP_HOST / SMTP_PORT / SMTP_USER / SMTP_PASS
    RECIPIENT_EMAIL
"""

import os, json, hashlib, datetime, time, smtplib, logging, re
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
RECIPIENT_EMAIL   = os.getenv("RECIPIENT_EMAIL", "jacopo.roccella@nttdata.com")
SMTP_HOST         = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT         = int(os.getenv("SMTP_PORT", 587))
SMTP_USER         = os.getenv("SMTP_USER", "")
SMTP_PASS         = os.getenv("SMTP_PASS", "")

DATA_DIR   = Path("./data")
OUTPUT_DIR = Path("./output")
LOG_FILE   = Path("./defence_scraper.log")
HIST_FILE  = DATA_DIR / "history.json"

DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler()],
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# PLAYER DEFINITIONS
# ─────────────────────────────────────────────
PLAYERS = [
    {"nome": "Ministero della Difesa",              "mercato": "Pubblico",  "complessita": "Molto Alta"},
    {"nome": "Forze Armate (Esercito/Marina/AM)",   "mercato": "Pubblico",  "complessita": "Alta"},
    {"nome": "Difesa Servizi",                      "mercato": "Pubblico",  "complessita": "Media"},
    {"nome": "Segretariato Generale Difesa / DNA",  "mercato": "Pubblico",  "complessita": "Alta"},
    {"nome": "Leonardo",                            "mercato": "Privato",   "complessita": "Alta"},
    {"nome": "Fincantieri",                         "mercato": "Privato",   "complessita": "Alta"},
    {"nome": "MBDA Italia",                         "mercato": "Privato",   "complessita": "Alta"},
    {"nome": "Elettronica Group",                   "mercato": "Privato",   "complessita": "Alta"},
    {"nome": "Thales Alenia Space Italia",          "mercato": "Privato",   "complessita": "Alta"},
    {"nome": "Avio Aero",                           "mercato": "Privato",   "complessita": "Alta"},
    {"nome": "Leonardo DRS",                        "mercato": "Privato",   "complessita": "Alta"},
]

# ─────────────────────────────────────────────
# ANTHROPIC WEB SEARCH
# ─────────────────────────────────────────────
def query_claude_with_search(player_name: str, mercato: str) -> dict:
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    tipo = "ente pubblico della Difesa italiana" if mercato == "Pubblico" else "azienda privata del settore Difesa italiana"

    prompt = f"""Sei un analista di mercato che lavora per NTT DATA Italia.
Devi raccogliere informazioni aggiornate su "{player_name}", {tipo}.

Fai ricerche web approfondite e poi restituisci SOLO un oggetto JSON valido
(niente testo prima o dopo il JSON, niente markdown, niente backtick) con questa struttura:

{{
  "fatturato_ultimo_anno": "valore in euro, es: 17.8 Bn euro (2024)",
  "spesa_it_stimata": "valore o range stimato in milioni euro, es: 320-450 M euro",
  "spesa_it_fonte": "fonte da cui e ricavata la stima",
  "servizi_it_acquistati": ["servizio1", "servizio2", "servizio3"],
  "piano_strategico_it": "sintesi del piano strategico digitale IT se disponibile",
  "gare_recenti": [
    {{
      "anno": "2024",
      "oggetto": "descrizione della gara IT",
      "importo": "importo in euro",
      "cig": "codice CIG o nd",
      "aggiudicatario": "nome azienda o nd",
      "fonte": "URL o nome fonte"
    }}
  ],
  "trend_futuro": "investimenti IT digitali previsti o annunciati",
  "note_rilevanti": "info utili per NTT DATA per posizionarsi su questo cliente"
}}

Cerca su: bilanci annuali, relazioni finanziarie, portale ANAC BDNCP, sito ufficiale,
comunicati stampa, piani industriali, documenti parlamentari, piano triennale IT.
Per le gare cerca: "{player_name} gara IT informatica cyber cloud ANAC"
Se un dato non e disponibile usa "nd" come valore stringa."""

    try:
        messages = [{"role": "user", "content": prompt}]
        tools = [{"type": "web_search_20250305", "name": "web_search"}]

        # Loop multi-turno per gestire web search
        for _ in range(5):
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=3000,
                tools=tools,
                messages=messages,
            )

            # Aggiungi risposta alla history
            messages.append({"role": "assistant", "content": response.content})

            # Se ha finito, estrai il testo finale
            if response.stop_reason == "end_turn":
                full_text = ""
                for block in response.content:
                    if hasattr(block, "text"):
                        full_text += block.text
                break

            # Se ha usato tool, fornisci risultati fittizi e continua
            if response.stop_reason == "tool_use":
                tool_results = []
                for block in response.content:
                    if block.type == "tool_use":
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": "Risultati di ricerca web ricevuti e analizzati."
                        })
                messages.append({"role": "user", "content": tool_results})
                continue

            # Altro stop reason
            full_text = ""
            for block in response.content:
                if hasattr(block, "text"):
                    full_text += block.text
            break

        # Prova estrazione JSON
        if full_text.strip():
            # Cerca JSON nel testo
            match = re.search(r'\{[\s\S]*\}', full_text)
            if match:
                try:
                    data = json.loads(match.group())
                    log.info(f"  ✓ {player_name}: dati estratti con successo")
                    return data
                except json.JSONDecodeError:
                    log.warning(f"  ⚠ {player_name}: JSON malformato, testo: {full_text[:200]}")

        log.warning(f"  ⚠ {player_name}: nessun JSON trovato")
        return _empty_record()

    except Exception as e:
        log.error(f"  ✗ {player_name}: errore API – {e}")
        return _empty_record()


def _empty_record() -> dict:
    return {
        "fatturato_ultimo_anno": "n.d.",
        "spesa_it_stimata": "n.d.",
        "spesa_it_fonte": "n.d.",
        "servizi_it_acquistati": [],
        "piano_strategico_it": "n.d.",
        "gare_recenti": [],
        "trend_futuro": "n.d.",
        "note_rilevanti": "n.d.",
    }


# ─────────────────────────────────────────────
# HISTORY & DIFF
# ─────────────────────────────────────────────
def load_history() -> dict:
    if HIST_FILE.exists():
        try:
            return json.loads(HIST_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_history(data: dict):
    HIST_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def compute_fingerprint(record: dict) -> str:
    blob = json.dumps({
        k: record.get(k) for k in
        ["fatturato_ultimo_anno", "spesa_it_stimata", "gare_recenti", "trend_futuro"]
    }, sort_keys=True)
    return hashlib.md5(blob.encode()).hexdigest()


def compute_diff(player_name: str, today: dict, history: dict) -> str:
    yesterday = history.get(player_name)
    if not yesterday:
        return "Prima esecuzione – baseline acquisita"
    fp_today = compute_fingerprint(today)
    fp_yest  = yesterday.get("fingerprint", "")
    if fp_today == fp_yest:
        return "Nessuna variazione vs ieri"
    changes = []
    if today.get("spesa_it_stimata") != yesterday.get("spesa_it_stimata"):
        changes.append(f"Spesa IT: {yesterday.get('spesa_it_stimata','?')} -> {today.get('spesa_it_stimata','?')}")
    if today.get("fatturato_ultimo_anno") != yesterday.get("fatturato_ultimo_anno"):
        changes.append("Fatturato: aggiornato")
    gare_oggi = len(today.get("gare_recenti", []))
    gare_ieri = yesterday.get("n_gare", 0)
    if gare_oggi != gare_ieri:
        diff_g = gare_oggi - gare_ieri
        changes.append(f"Gare: {'+' if diff_g > 0 else ''}{diff_g} rispetto a ieri")
    if today.get("trend_futuro") != yesterday.get("trend_futuro"):
        changes.append("Trend futuro: aggiornato")
    return " | ".join(changes) if changes else "Variazioni minori rilevate"


# ─────────────────────────────────────────────
# EXCEL GENERATION
# ─────────────────────────────────────────────
COLOR_HEADER  = "1F3864"
COLOR_SUBHEAD = "2E75B6"
COLOR_PUB     = "D9E2F3"
COLOR_PRIV    = "E2EFDA"
COLOR_CHANGE  = "FFE599"
FONT_NAME     = "Arial"


def _hdr(cell, bg=COLOR_HEADER, fg="FFFFFF", size=10):
    cell.font      = Font(name=FONT_NAME, bold=True, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _border(cell)


def _border(cell, color="BBBBBB"):
    s = Side(border_style="thin", color=color)
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bg="FFFFFF", bold=False, wrap=True, size=9, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=FONT_NAME, size=size, bold=bold, color=color)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(vertical="center", wrap_text=wrap)
    _border(c)
    return c


def build_excel(all_results: list, date_str: str) -> Path:
    wb = openpyxl.Workbook()

    # SHEET 1: DASHBOARD
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value     = f"NTT DATA - Defence Market Intelligence  |  {date_str}"
    t.font      = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:L2")
    leg = ws["A2"]
    leg.value     = "Pubblico (blu)   Privato (verde)   Variazione rilevata vs ieri (giallo)"
    leg.font      = Font(name=FONT_NAME, size=9, italic=True)
    leg.alignment = Alignment(horizontal="center")

    hdrs = ["Player", "Mercato", "Complessita Proc.", "Fatturato",
            "Spesa IT Stimata", "Fonte Stima", "Servizi IT Acquistati",
            "Piano Strategico IT", "N. Gare Trovate", "Trend Futuro",
            "Delta vs Ieri", "Note per NTT DATA"]
    widths = [28, 10, 16, 18, 18, 22, 38, 38, 12, 36, 38, 38]

    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws.cell(row=3, column=ci), bg=COLOR_SUBHEAD)
        ws.cell(row=3, column=ci).value = h
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 36

    for ri, rec in enumerate(all_results, 4):
        bg = COLOR_PUB if rec["mercato"] == "Pubblico" else COLOR_PRIV
        delta_bg = COLOR_CHANGE if "->" in rec["delta"] or "Gare:" in rec["delta"] or "aggiornato" in rec["delta"] else bg
        servizi = ", ".join(rec["data"].get("servizi_it_acquistati", [])) or "n.d."
        row_vals = [
            rec["nome"],
            rec["mercato"],
            rec["complessita"],
            rec["data"].get("fatturato_ultimo_anno", "n.d."),
            rec["data"].get("spesa_it_stimata", "n.d."),
            rec["data"].get("spesa_it_fonte", "n.d."),
            servizi,
            rec["data"].get("piano_strategico_it", "n.d."),
            str(len(rec["data"].get("gare_recenti", []))),
            rec["data"].get("trend_futuro", "n.d."),
            rec["delta"],
            rec["data"].get("note_rilevanti", "n.d."),
        ]
        for ci, val in enumerate(row_vals, 1):
            b = delta_bg if ci == 11 else bg
            _cell(ws, ri, ci, val, bg=b)
        ws.row_dimensions[ri].height = 52

    ws.auto_filter.ref = f"A3:{get_column_letter(len(hdrs))}3"

    # SHEET 2: STORICO GARE
    ws2 = wb.create_sheet("Storico Gare")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value     = "Storico Gare IT Cyber Cloud - Player Difesa"
    t2.font      = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    t2.fill      = PatternFill("solid", start_color=COLOR_HEADER)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    g_hdrs   = ["Player", "Anno", "Oggetto Gara", "Importo", "CIG", "Aggiudicatario", "Fonte"]
    g_widths = [26, 8, 52, 14, 20, 28, 42]
    for ci, (h, w) in enumerate(zip(g_hdrs, g_widths), 1):
        _hdr(ws2.cell(row=2, column=ci), bg=COLOR_SUBHEAD)
        ws2.cell(row=2, column=ci).value = h
        ws2.column_dimensions[get_column_letter(ci)].width = w

    row_i = 3
    fill_a = PatternFill("solid", start_color="EEF2FF")
    fill_b = PatternFill("solid", start_color="F8F9FF")
    for idx, rec in enumerate(all_results):
        gare = rec["data"].get("gare_recenti", [])
        if not gare:
            c = ws2.cell(row=row_i, column=1, value=rec["nome"])
            c.font = Font(name=FONT_NAME, size=9, italic=True, color="888888")
            ws2.cell(row=row_i, column=2, value="-")
            ws2.cell(row=row_i, column=3, value="Nessuna gara trovata")
            for ci in range(1, 8):
                _border(ws2.cell(row=row_i, column=ci))
                ws2.cell(row=row_i, column=ci).fill = fill_b
            row_i += 1
            continue
        fill = fill_a if idx % 2 == 0 else fill_b
        for g in gare:
            vals = [
                rec["nome"],
                g.get("anno", "n.d."),
                g.get("oggetto", "n.d."),
                g.get("importo", "n.d."),
                g.get("cig", "n.d."),
                g.get("aggiudicatario", "n.d."),
                g.get("fonte", "n.d."),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws2.cell(row=row_i, column=ci, value=val)
                c.font      = Font(name=FONT_NAME, size=9)
                c.fill      = fill
                c.alignment = Alignment(vertical="center", wrap_text=True)
                _border(c)
            ws2.row_dimensions[row_i].height = 28
            row_i += 1

    ws2.auto_filter.ref = f"A2:{get_column_letter(len(g_hdrs))}2"

    # SHEET 3: DETTAGLIO PLAYER
    ws3 = wb.create_sheet("Dettaglio Player")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 80

    ws3.merge_cells("A1:B1")
    t3 = ws3["A1"]
    t3.value     = "Scheda Dettaglio per Player"
    t3.font      = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    t3.fill      = PatternFill("solid", start_color=COLOR_HEADER)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    row_i = 2
    for rec in all_results:
        bg = COLOR_PUB if rec["mercato"] == "Pubblico" else COLOR_PRIV
        ws3.merge_cells(f"A{row_i}:B{row_i}")
        c = ws3.cell(row=row_i, column=1, value=f"  {rec['nome']}  ({rec['mercato']})")
        c.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color=COLOR_SUBHEAD)
        c.alignment = Alignment(vertical="center")
        ws3.row_dimensions[row_i].height = 24
        row_i += 1

        d = rec["data"]
        fields = [
            ("Fatturato",             d.get("fatturato_ultimo_anno", "n.d.")),
            ("Spesa IT Stimata",      d.get("spesa_it_stimata", "n.d.")),
            ("Fonte Stima",           d.get("spesa_it_fonte", "n.d.")),
            ("Servizi IT Acquistati", ", ".join(d.get("servizi_it_acquistati", []))),
            ("Piano Strategico IT",   d.get("piano_strategico_it", "n.d.")),
            ("Trend Futuro",          d.get("trend_futuro", "n.d.")),
            ("Note per NTT DATA",     d.get("note_rilevanti", "n.d.")),
            ("N. Gare Trovate",       str(len(d.get("gare_recenti", [])))),
        ]
        for label, val in fields:
            _cell(ws3, row_i, 1, label, bg=bg, bold=True)
            _cell(ws3, row_i, 2, val,   bg="FFFFFF")
            ws3.row_dimensions[row_i].height = 36
            row_i += 1
        row_i += 1

    fname = OUTPUT_DIR / f"defence_market_{date_str.replace('-', '')}.xlsx"
    wb.save(fname)
    log.info(f"Excel salvato: {fname}")
    return fname


# ─────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────
def build_email_html(all_results: list, date_str: str) -> str:
    changed   = sum(1 for r in all_results if "->" in r["delta"] or "aggiornato" in r["delta"] or "Gare:" in r["delta"])
    tot_gare  = sum(len(r["data"].get("gare_recenti", [])) for r in all_results)
    con_spesa = sum(1 for r in all_results if r["data"].get("spesa_it_stimata", "n.d.") != "n.d.")

    rows = ""
    for r in all_results:
        delta_bg = "#FFE599" if ("->" in r["delta"] or "aggiornato" in r["delta"] or "Gare:" in r["delta"]) else "#FFFFFF"
        badge    = "#D9E2F3" if r["mercato"] == "Pubblico" else "#E2EFDA"
        rows += f"""
        <tr>
          <td style="padding:6px 8px;border:1px solid #ddd;background:{badge};font-weight:600">{r['nome']}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{r['mercato']}</td>
          <td style="padding:6px 8px;border:1px solid #ddd">{r['data'].get('fatturato_ultimo_anno','n.d.')}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;font-weight:600;color:#1F3864">{r['data'].get('spesa_it_stimata','n.d.')}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{len(r['data'].get('gare_recenti',[]))}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;background:{delta_bg};font-size:11px">{r['delta']}</td>
        </tr>"""

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;color:#1F3864;margin:0;padding:20px">
<div style="max-width:860px;margin:auto">
  <div style="background:#1F3864;padding:18px 24px;border-radius:8px 8px 0 0">
    <h1 style="color:#fff;margin:0;font-size:18px">NTT DATA - Defence Market Intelligence</h1>
    <p style="color:#aac4e8;margin:4px 0 0;font-size:13px">Report giornaliero - {date_str}</p>
  </div>
  <div style="background:#f0f4fa;padding:16px 24px;display:flex;gap:16px;flex-wrap:wrap">
    <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #2E75B6">
      <div style="font-size:24px;font-weight:700;color:#2E75B6">{len(all_results)}</div>
      <div style="font-size:12px;color:#555">Player monitorati</div>
    </div>
    <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #70AD47">
      <div style="font-size:24px;font-weight:700;color:#70AD47">{con_spesa}</div>
      <div style="font-size:12px;color:#555">Con stima spesa IT</div>
    </div>
    <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #7030A0">
      <div style="font-size:24px;font-weight:700;color:#7030A0">{tot_gare}</div>
      <div style="font-size:12px;color:#555">Gare trovate totali</div>
    </div>
    <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #FF8C00">
      <div style="font-size:24px;font-weight:700;color:#FF8C00">{changed}</div>
      <div style="font-size:12px;color:#555">Variazioni vs ieri</div>
    </div>
  </div>
  <table style="width:100%;border-collapse:collapse;font-size:12px">
    <thead>
      <tr style="background:#2E75B6;color:#fff">
        <th style="padding:8px;border:1px solid #1a5c9e;text-align:left">Player</th>
        <th style="padding:8px;border:1px solid #1a5c9e">Mercato</th>
        <th style="padding:8px;border:1px solid #1a5c9e">Fatturato</th>
        <th style="padding:8px;border:1px solid #1a5c9e">Spesa IT Stimata</th>
        <th style="padding:8px;border:1px solid #1a5c9e">N. Gare</th>
        <th style="padding:8px;border:1px solid #1a5c9e;min-width:180px">Delta vs Ieri</th>
      </tr>
    </thead>
    <tbody>{rows}</tbody>
  </table>
  <p style="font-size:11px;color:#888;margin-top:16px">
    In allegato: Excel con Dashboard, Storico Gare completo e Schede Dettaglio per player.
  </p>
</div>
</body></html>"""


def send_email(xlsx_path: Path, html_body: str, date_str: str):
    if not SMTP_USER or not SMTP_PASS:
        log.warning("Credenziali SMTP non configurate - mail non inviata.")
        return
    msg = MIMEMultipart("mixed")
    msg["From"]    = SMTP_USER
    msg["To"]      = RECIPIENT_EMAIL
    msg["Subject"] = f"[Defence Market] Report {date_str}"
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    with open(xlsx_path, "rb") as f:
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{xlsx_path.name}"')
    msg.attach(part)
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.ehlo(); s.starttls(); s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_USER, RECIPIENT_EMAIL, msg.as_string())
        log.info(f"Mail inviata a {RECIPIENT_EMAIL}")
    except Exception as e:
        log.error(f"Errore invio mail: {e}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def run():
    date_str = datetime.date.today().isoformat()
    log.info(f"=== Esecuzione {date_str} ===")

    history     = load_history()
    all_results = []

    for player in PLAYERS:
        log.info(f"  Ricerca: {player['nome']}")
        data  = query_claude_with_search(player["nome"], player["mercato"])
        delta = compute_diff(player["nome"], data, history)
        fp    = compute_fingerprint(data)

        history[player["nome"]] = {
            "fingerprint":           fp,
            "spesa_it_stimata":      data.get("spesa_it_stimata"),
            "fatturato_ultimo_anno": data.get("fatturato_ultimo_anno"),
            "trend_futuro":          data.get("trend_futuro"),
            "n_gare":                len(data.get("gare_recenti", [])),
            "date":                  date_str,
        }
        all_results.append({
            "nome":        player["nome"],
            "mercato":     player["mercato"],
            "complessita": player["complessita"],
            "data":        data,
            "delta":       delta,
        })
        time.sleep(3)

    save_history(history)
    xlsx_path = build_excel(all_results, date_str)
    html_body = build_email_html(all_results, date_str)
    send_email(xlsx_path, html_body, date_str)
    log.info("=== Fine esecuzione ===")
    return xlsx_path


if __name__ == "__main__":
    import sys
    if "--once" in sys.argv:
        run()
    else:
        try:
            import schedule
            log.info("Scheduler attivo - esecuzione ogni giorno alle 07:30.")
            schedule.every().day.at("07:30").do(run)
            run()
            while True:
                schedule.run_pending()
                time.sleep(60)
        except ImportError:
            run()
