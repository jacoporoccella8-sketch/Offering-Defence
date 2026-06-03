"""
Defence Market Analysis – NTT DATA Italy
Daily scraper: raccoglie dati sui player del segmento Difesa,
genera un Excel di sintesi e invia una mail di aggiornamento.

Requisiti:
    pip install requests beautifulsoup4 openpyxl schedule python-dotenv

Variabili d'ambiente (file .env):
    ANTHROPIC_API_KEY   – chiave API Anthropic (usata per parsing intelligente)
    SMTP_HOST           – es. smtp.office365.com
    SMTP_PORT           – es. 587
    SMTP_USER           – indirizzo mittente
    SMTP_PASS           – password / app-password
    RECIPIENT_EMAIL     – jacopo.roccella@nttdata.com
"""

import os, json, re, smtplib, hashlib, datetime, time, logging
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
RECIPIENT_EMAIL = os.getenv("RECIPIENT_EMAIL", "jacopo.roccella@nttdata.com")
SMTP_HOST       = os.getenv("SMTP_HOST", "smtp.office365.com")
SMTP_PORT       = int(os.getenv("SMTP_PORT", 587))
SMTP_USER       = os.getenv("SMTP_USER", "")
SMTP_PASS       = os.getenv("SMTP_PASS", "")

DATA_DIR   = Path("./data")
OUTPUT_DIR = Path("./output")
LOG_FILE   = Path("./defence_scraper.log")
HIST_FILE  = DATA_DIR / "history.json"

DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler()],
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# PLAYER DEFINITIONS
# ─────────────────────────────────────────────
PLAYERS = [
    {
        "nome":     "Ministero della Difesa",
        "mercato":  "Pubblico",
        "ruolo":    "Buyer centrale – definisce fabbisogni e indirizzi",
        "complessita_procurement": "Molto Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.difesa.it/content/statoprevisionespesa/26765.html",
            "https://www.difesa.it/assets/allegati/3754/piano_di_analisi_e_valutazione_della_spesa_2025-2027_del_ministero_della_difesa.pdf",
            "https://www.difesa.it/primopiano/il-ministero-della-difesa-sceglie-il-cloud-di-polo-strategico-nazionale/53268.html",
        ],
        "servizi_it": ["Cloud (PSN)", "Cybersecurity", "Data", "Legacy migration",
                       "Applicativi", "Governance IT"],
    },
    {
        "nome":     "Forze Armate",
        "mercato":  "Pubblico",
        "ruolo":    "Utenti/committenti operativi",
        "complessita_procurement": "Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.senato.it/service/PDF/PDFServer/BGT/001484480.pdf",
            "https://pubblicitalegale.anticorruzione.it/bdncp",
        ],
        "servizi_it": ["Infrastrutture", "Reti", "SOC", "Collaboration",
                       "Mobile", "Training", "Mission-support systems"],
    },
    {
        "nome":     "Difesa Servizi",
        "mercato":  "Pubblico",
        "ruolo":    "In-house del Ministero – valorizza asset e gestisce iniziative",
        "complessita_procurement": "Media",
        "fatturato_bn": None,
        "fonti": [
            "https://www.difesaservizi.it/gare",
        ],
        "servizi_it": ["Servizi digitali", "Comunicazione", "Piattaforme",
                       "Procurement indiretto"],
    },
    {
        "nome":     "Esercito / Marina / Aeronautica",
        "mercato":  "Pubblico",
        "ruolo":    "Branch operativi con esigenze specifiche",
        "complessita_procurement": "Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.esercito.difesa.it/comunicazione/bandi-di-gara-esercito/bandi-di-gara/126748.html",
            "https://www.marina.difesa.it/Pagine/default.aspx",
            "https://www.aeronautica.difesa.it/",
        ],
        "servizi_it": ["Postazioni", "Reti", "SOC", "Analytics",
                       "Manutenzione", "Training", "Collaboration"],
    },
    {
        "nome":     "Segretariato Generale Difesa / DNA",
        "mercato":  "Pubblico",
        "ruolo":    "Hub amministrativo-tecnico e programmatico",
        "complessita_procurement": "Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.difesa.it/assets/allegati/3754/piano_di_analisi_e_valutazione_della_spesa_2025-2027_del_ministero_della_difesa.pdf",
            "https://pubblicitalegale.anticorruzione.it/bandi",
        ],
        "servizi_it": ["Program management", "Procurement support",
                       "Document management", "Compliance", "Architecture"],
    },
    {
        "nome":     "Leonardo",
        "mercato":  "Privato",
        "ruolo":    "Prime contractor e integratore strategico",
        "complessita_procurement": "Alta",
        "fatturato_bn": 17.8,
        "fonti": [
            "https://www.leonardo.com/en/investors/results-and-reports",
            "https://cybersecurity.leonardo.com/en/digitalisation",
            "https://www.leonardo.com/documents/15646808/0/2024+Integrated+Annual+Report.pdf",
        ],
        "servizi_it": ["Cyber", "Cloud", "Data", "Software engineering",
                       "Digital HMI", "Consulenza specialistica"],
    },
    {
        "nome":     "Fincantieri",
        "mercato":  "Privato",
        "ruolo":    "Prime contractor navale e industriale",
        "complessita_procurement": "Alta",
        "fatturato_bn": 9.19,
        "fonti": [
            "https://www.fincantieri.com/it/investor-relations/dati-documenti-e-financial-highlights/dati-finanziari",
            "https://www.fincantieri.com/en/business/products/systems--components-e-infrastructures/cybersecurity",
        ],
        "servizi_it": ["OT security", "e-Procurement", "PLM",
                       "Digital shipyard", "Cyber resilience", "Data"],
    },
    {
        "nome":     "MBDA Italia",
        "mercato":  "Privato",
        "ruolo":    "Missile/defence systems – programmi complessi",
        "complessita_procurement": "Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.mbda-systems.com/sites/mbda/files/2025-06/mbda_sustainability-report-2024_webfriendly.pdf",
        ],
        "servizi_it": ["PLM", "Engineering collaboration", "Security",
                       "Testing", "Systems integration"],
    },
    {
        "nome":     "Elettronica Group",
        "mercato":  "Privato",
        "ruolo":    "Elettronica per difesa ed EW",
        "complessita_procurement": "Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.eltgroup.net/",
        ],
        "servizi_it": ["Secure engineering", "Data", "Manufacturing digitalization",
                       "Cyber OT/IT", "Compliance"],
    },
    {
        "nome":     "Thales Alenia Space Italia",
        "mercato":  "Privato",
        "ruolo":    "Spazio dual-use – mission-critical",
        "complessita_procurement": "Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.thalesaleniaspace.com/en",
        ],
        "servizi_it": ["Cloud", "Data", "AI", "Cyber",
                       "Digital engineering", "Simulazione"],
    },
    {
        "nome":     "Avio Aero",
        "mercato":  "Privato",
        "ruolo":    "Industrial aerospace/defence supply chain",
        "complessita_procurement": "Alta",
        "fatturato_bn": None,
        "fonti": [
            "https://www.avioaero.com/",
        ],
        "servizi_it": ["ERP", "MES", "PLM", "Analytics",
                       "Quality", "Supply chain", "Cyber OT/IT"],
    },
]

# Fonti per gare/procurement
PROCUREMENT_SOURCES = [
    "https://pubblicitalegale.anticorruzione.it/bdncp",
    "https://www.difesaservizi.it/gare",
]

# ─────────────────────────────────────────────
# SCRAPING
# ─────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; NTTDataDefenceBot/1.0; "
        "+https://www.nttdata.com)"
    )
}

def fetch_page(url: str, timeout: int = 15) -> str | None:
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r.text
    except Exception as e:
        log.warning(f"Fetch failed {url}: {e}")
        return None


def extract_text_snippet(html: str, max_chars: int = 800) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    text = re.sub(r"\s+", " ", soup.get_text(separator=" ")).strip()
    return text[:max_chars]


def extract_financial_keywords(text: str) -> dict:
    """Cerca keyword finanziarie / IT nel testo di una pagina."""
    keywords = {
        "cyber": bool(re.search(r"cyber|sicurezza\s+informatica|cybersecurity", text, re.I)),
        "cloud": bool(re.search(r"cloud|nuvola\s+informatica", text, re.I)),
        "data_ai": bool(re.search(r"\bdata\b|intelligenza\s+artificiale|machine\s+learning|AI\b", text, re.I)),
        "digital_engineering": bool(re.search(r"digital\s+engineering|ingegneria\s+digitale|PLM|MES|ERP", text, re.I)),
        "gara_appalto": bool(re.search(r"gara|appalto|bando|affidamento|CIG|base\s+d.asta", text, re.I)),
        "budget_spesa": bool(re.search(r"budget|milion|miliard|mln|mrd|spesa\s+IT|investiment", text, re.I)),
    }
    # Prova a estrarre cifre in €M / €Bn
    amounts = re.findall(r"([\d\.,]+)\s*(?:milioni|miliardi|mln|mrd|bn|B€|M€|\bM\b|\bB\b)", text, re.I)
    keywords["amounts_found"] = amounts[:5]
    return keywords


def scrape_player(player: dict) -> dict:
    log.info(f"  Scraping: {player['nome']}")
    results = {
        "snippets": [],
        "keywords_aggregate": {
            "cyber": False, "cloud": False, "data_ai": False,
            "digital_engineering": False, "gara_appalto": False, "budget_spesa": False,
            "amounts_found": [],
        },
        "fonti_ok": 0,
        "fonti_ko": 0,
        "last_update": datetime.datetime.now().isoformat(),
    }
    for url in player["fonti"]:
        html = fetch_page(url)
        if html:
            snippet = extract_text_snippet(html)
            kw = extract_financial_keywords(snippet)
            results["snippets"].append({"url": url, "snippet": snippet[:300]})
            results["fonti_ok"] += 1
            for k in ["cyber", "cloud", "data_ai", "digital_engineering",
                      "gara_appalto", "budget_spesa"]:
                if kw[k]:
                    results["keywords_aggregate"][k] = True
            results["keywords_aggregate"]["amounts_found"].extend(kw["amounts_found"])
        else:
            results["fonti_ko"] += 1
    return results


# ─────────────────────────────────────────────
# HISTORY & DIFF
# ─────────────────────────────────────────────
def load_history() -> dict:
    if HIST_FILE.exists():
        try:
            return json.loads(HIST_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def save_history(data: dict):
    HIST_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def compute_fingerprint(scrape_result: dict) -> str:
    blob = json.dumps(scrape_result.get("keywords_aggregate", {}), sort_keys=True)
    return hashlib.md5(blob.encode()).hexdigest()


def compute_diff(player_name: str, today_result: dict, history: dict) -> str:
    """Restituisce una stringa human-readable del delta rispetto a ieri."""
    yesterday = history.get(player_name)
    if not yesterday:
        return "Prima esecuzione – nessun confronto disponibile"

    fp_today = compute_fingerprint(today_result)
    fp_yesterday = yesterday.get("fingerprint", "")
    if fp_today == fp_yesterday:
        return "Nessuna variazione rispetto a ieri"

    changes = []
    kw_today = today_result.get("keywords_aggregate", {})
    kw_yest  = yesterday.get("keywords_aggregate", {})
    for k in ["cyber", "cloud", "data_ai", "digital_engineering",
              "gara_appalto", "budget_spesa"]:
        if kw_today.get(k) != kw_yest.get(k):
            stato = "rilevato" if kw_today.get(k) else "scomparso"
            changes.append(f"'{k}' {stato}")

    amounts_new = set(kw_today.get("amounts_found", []))
    amounts_old = set(kw_yest.get("amounts_found", []))
    new_a = amounts_new - amounts_old
    if new_a:
        changes.append(f"Nuovi importi: {', '.join(new_a)}")

    ok_d = today_result.get("fonti_ok", 0) - yesterday.get("fonti_ok", 0)
    if ok_d != 0:
        changes.append(f"Fonti raggiungibili: {'+' if ok_d > 0 else ''}{ok_d}")

    return " | ".join(changes) if changes else "Modifiche minori non classificate"


# ─────────────────────────────────────────────
# EXCEL GENERATION
# ─────────────────────────────────────────────
COLOR_HEADER  = "1F3864"   # Navy NTT
COLOR_SUBHEAD = "2E75B6"
COLOR_PUB     = "D9E2F3"
COLOR_PRIV    = "E2EFDA"
COLOR_CHANGE  = "FFE599"   # giallo per delta
COLOR_OK      = "C6EFCE"
COLOR_KO      = "FFC7CE"
FONT_NAME     = "Arial"


def _hdr_style(cell, bg=COLOR_HEADER, fg="FFFFFF", bold=True, size=10):
    cell.font      = Font(name=FONT_NAME, bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border_all(cell, color="AAAAAA"):
    s = Side(border_style="thin", color=color)
    cell.border = Border(left=s, right=s, top=s, bottom=s)


def build_excel(all_results: list[dict], date_str: str) -> Path:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Dashboard ────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Title
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"NTT DATA – Market Analysis Difesa  |  Aggiornato: {date_str}"
    title_cell.font  = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", start_color=COLOR_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Legend row
    ws.merge_cells("A2:N2")
    leg = ws["A2"]
    leg.value = "🟦 Pubblico   🟩 Privato   🟨 Variazione vs ieri"
    leg.font  = Font(name=FONT_NAME, size=9, italic=True)
    leg.alignment = Alignment(horizontal="center")

    # Column headers
    headers = [
        "Player", "Mercato", "Ruolo", "Complessità Proc.",
        "Fatturato (Bn€)", "Cyber", "Cloud", "Data/AI",
        "Digital Eng.", "Gare/Appalti", "Budget/Spesa",
        "Fonti OK/TOT", "Delta vs ieri", "Ultimo aggiornamento",
    ]
    col_widths = [28, 10, 38, 16, 14, 8, 8, 8, 12, 12, 12, 12, 42, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        _hdr_style(c, bg=COLOR_SUBHEAD)
        _border_all(c)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 36

    # Data rows
    for ri, rec in enumerate(all_results, 4):
        bg = COLOR_PUB if rec["mercato"] == "Pubblico" else COLOR_PRIV
        kw = rec.get("keywords_aggregate", {})

        def yn(val):
            return "✔" if val else "–"

        fonti_str = f"{rec.get('fonti_ok',0)}/{rec.get('fonti_ok',0)+rec.get('fonti_ko',0)}"
        row_vals = [
            rec["nome"],
            rec["mercato"],
            rec["ruolo"],
            rec["complessita"],
            rec["fatturato_bn"] if rec["fatturato_bn"] else "n.d.",
            yn(kw.get("cyber")),
            yn(kw.get("cloud")),
            yn(kw.get("data_ai")),
            yn(kw.get("digital_engineering")),
            yn(kw.get("gara_appalto")),
            yn(kw.get("budget_spesa")),
            fonti_str,
            rec["delta"],
            rec.get("last_update", "")[:16].replace("T", " "),
        ]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name=FONT_NAME, size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            _border_all(c)
            # background
            if ci in (6, 7, 8, 9, 10, 11):  # keyword booleans
                if val == "✔":
                    c.fill = PatternFill("solid", start_color=COLOR_OK)
                else:
                    c.fill = PatternFill("solid", start_color="F4F4F4")
            elif ci == 13 and "variazione" in rec["delta"].lower().replace("nessuna", ""):
                c.fill = PatternFill("solid", start_color=COLOR_CHANGE)
            else:
                c.fill = PatternFill("solid", start_color=bg)

        ws.row_dimensions[ri].height = 30

    # Auto filter
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"

    # ── Sheet 2: Gare & Procurement ───────────
    ws2 = wb.create_sheet("Gare & Procurement")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]
    t2.value = "Gare e Procurement Difesa – Monitoraggio"
    t2.font  = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    t2.fill  = PatternFill("solid", start_color=COLOR_HEADER)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    g_headers = ["Player / Ente", "URL Fonte", "Oggetto / Keyword", "Gara Rilevata",
                 "Importo Trovato", "Data Rilevazione"]
    g_widths  = [26, 48, 36, 14, 18, 18]
    for ci, (h, w) in enumerate(zip(g_headers, g_widths), 1):
        c = ws2.cell(row=2, column=ci, value=h)
        _hdr_style(c, bg=COLOR_SUBHEAD)
        _border_all(c)
        ws2.column_dimensions[get_column_letter(ci)].width = w

    fill_light = PatternFill("solid", start_color="EEF2FF")
    for ri, rec in enumerate(all_results, 3):
        kw = rec.get("keywords_aggregate", {})
        importi = ", ".join(kw.get("amounts_found", [])) or "–"
        row_v = [
            rec["nome"],
            rec["fonti_principali"],
            ", ".join(rec.get("servizi_it", [])),
            "✔" if kw.get("gara_appalto") else "–",
            importi,
            rec.get("last_update", "")[:16].replace("T", " "),
        ]
        for ci, val in enumerate(row_v, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font      = Font(name=FONT_NAME, size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            _border_all(c)
            c.fill = fill_light
        ws2.row_dimensions[ri].height = 28

    # ── Sheet 3: Fonti & Metodologia ─────────
    ws3 = wb.create_sheet("Fonti & Metodologia")
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 70
    ws3.column_dimensions["C"].width = 40

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = "Fonti di ricerca per player – Metodologia 8 piste"
    t3.font  = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    t3.fill  = PatternFill("solid", start_color=COLOR_HEADER)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    for ci, h in enumerate(["Player", "URL Fonte", "Tipo pista"], 1):
        c = ws3.cell(row=2, column=ci, value=h)
        _hdr_style(c, bg=COLOR_SUBHEAD)

    row_i = 3
    for rec in all_results:
        for url in rec.get("fonti_list", []):
            pista = "Procurement/Gare" if any(k in url for k in ["anac", "difesaservizi", "gare", "bdncp"]) \
                else "Industriale/Bilancio" if any(k in url for k in ["leonardo", "fincantieri", "mbda", "eltgroup", "avioaero", "thalesalenia"]) \
                else "Normativa/Pubblica"
            for ci, val in enumerate([rec["nome"], url, pista], 1):
                c = ws3.cell(row=row_i, column=ci, value=val)
                c.font      = Font(name=FONT_NAME, size=9)
                c.alignment = Alignment(vertical="center")
                _border_all(c)
                if ci == 2:
                    c.hyperlink = url
                    c.font = Font(name=FONT_NAME, size=9, color="0563C1", underline="single")
            row_i += 1

    # ── Save ──────────────────────────────────
    fname = OUTPUT_DIR / f"defence_market_{date_str.replace('-', '')}.xlsx"
    wb.save(fname)
    log.info(f"Excel salvato: {fname}")
    return fname


# ─────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────
def build_email_html(all_results: list[dict], date_str: str) -> str:
    rows_html = ""
    changed = sum(
        1 for r in all_results
        if "variazione" in r["delta"].lower() or "nuovo" in r["delta"].lower()
    )
    ok_pct = round(
        100 * sum(r.get("fonti_ok", 0) for r in all_results) /
        max(sum(r.get("fonti_ok", 0) + r.get("fonti_ko", 0) for r in all_results), 1)
    )

    for r in all_results:
        badge_color = "#c6efce" if r["mercato"] == "Pubblico" else "#e2efda"
        delta_bg    = "#FFE599" if (
            "variazione" in r["delta"].lower() or "nuovo" in r["delta"].lower()
        ) else "#FFFFFF"
        kw = r.get("keywords_aggregate", {})
        def dot(v):
            return "🟢" if v else "⚪"
        rows_html += f"""
        <tr>
          <td style="padding:6px 8px;border:1px solid #ddd;background:{badge_color};font-weight:600">{r['nome']}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{r['mercato']}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{dot(kw.get('cyber'))}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{dot(kw.get('cloud'))}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{dot(kw.get('data_ai'))}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{dot(kw.get('gara_appalto'))}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;text-align:center">{r.get('fonti_ok',0)}/{r.get('fonti_ok',0)+r.get('fonti_ko',0)}</td>
          <td style="padding:6px 8px;border:1px solid #ddd;background:{delta_bg};font-size:11px">{r['delta']}</td>
        </tr>"""

    return f"""
<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;color:#1F3864;margin:0;padding:20px">
  <div style="max-width:900px;margin:auto">
    <div style="background:#1F3864;padding:18px 24px;border-radius:8px 8px 0 0">
      <h1 style="color:#fff;margin:0;font-size:18px">
        NTT DATA – Defence Market Analysis
      </h1>
      <p style="color:#aac4e8;margin:4px 0 0;font-size:13px">
        Report giornaliero · {date_str}
      </p>
    </div>

    <!-- Summary boxes -->
    <div style="background:#f0f4fa;padding:16px 24px;display:flex;gap:20px;flex-wrap:wrap">
      <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #2E75B6;min-width:130px">
        <div style="font-size:24px;font-weight:700;color:#2E75B6">{len(all_results)}</div>
        <div style="font-size:12px;color:#555">Player monitorati</div>
      </div>
      <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #FF8C00;min-width:130px">
        <div style="font-size:24px;font-weight:700;color:#FF8C00">{changed}</div>
        <div style="font-size:12px;color:#555">Con variazioni vs ieri</div>
      </div>
      <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #70AD47;min-width:130px">
        <div style="font-size:24px;font-weight:700;color:#70AD47">{ok_pct}%</div>
        <div style="font-size:12px;color:#555">Fonti raggiungibili</div>
      </div>
      <div style="background:#fff;border-radius:6px;padding:12px 20px;border-left:4px solid #7030A0;min-width:130px">
        <div style="font-size:24px;font-weight:700;color:#7030A0">
          {sum(1 for r in all_results if r.get('keywords_aggregate',{}).get('gara_appalto'))}
        </div>
        <div style="font-size:12px;color:#555">Con gare rilevate</div>
      </div>
    </div>

    <!-- Main table -->
    <table style="width:100%;border-collapse:collapse;font-size:12px;margin-top:0">
      <thead>
        <tr style="background:#2E75B6;color:#fff">
          <th style="padding:8px;border:1px solid #1a5c9e;text-align:left">Player</th>
          <th style="padding:8px;border:1px solid #1a5c9e">Mercato</th>
          <th style="padding:8px;border:1px solid #1a5c9e">Cyber</th>
          <th style="padding:8px;border:1px solid #1a5c9e">Cloud</th>
          <th style="padding:8px;border:1px solid #1a5c9e">Data/AI</th>
          <th style="padding:8px;border:1px solid #1a5c9e">Gare</th>
          <th style="padding:8px;border:1px solid #1a5c9e">Fonti</th>
          <th style="padding:8px;border:1px solid #1a5c9e;min-width:200px">Delta vs ieri</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>

    <p style="font-size:11px;color:#888;margin-top:16px">
      In allegato trovi il file Excel con il dettaglio completo (Dashboard, Gare &amp; Procurement, Fonti &amp; Metodologia).<br>
      🟡 Celle gialle nella colonna "Delta" = variazione rilevata rispetto all'esecuzione precedente.
    </p>
    <p style="font-size:11px;color:#bbb">
      Script autogenerato da <strong>NTT DATA Defence Market Scraper</strong> – esecuzione automatica quotidiana.
    </p>
  </div>
</body></html>"""


def send_email(xlsx_path: Path, html_body: str, date_str: str):
    if not SMTP_USER or not SMTP_PASS:
        log.warning("Credenziali SMTP non configurate – mail non inviata.")
        return

    msg = MIMEMultipart("mixed")
    msg["From"]    = SMTP_USER
    msg["To"]      = RECIPIENT_EMAIL
    msg["Subject"] = f"[Defence Market] Report giornaliero – {date_str}"

    msg.attach(MIMEText(html_body, "html", "utf-8"))

    with open(xlsx_path, "rb") as f:
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{xlsx_path.name}"')
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, RECIPIENT_EMAIL, msg.as_string())
        log.info(f"Mail inviata a {RECIPIENT_EMAIL}")
    except Exception as e:
        log.error(f"Errore invio mail: {e}")


# ─────────────────────────────────────────────
# MAIN RUN
# ─────────────────────────────────────────────
def run():
    date_str = datetime.date.today().isoformat()
    log.info(f"=== Esecuzione {date_str} ===")

    history = load_history()
    all_results = []

    for player in PLAYERS:
        scrape = scrape_player(player)
        fp     = compute_fingerprint(scrape)
        delta  = compute_diff(player["nome"], scrape, history)

        # aggiorna storia
        history[player["nome"]] = {
            "fingerprint":        fp,
            "keywords_aggregate": scrape["keywords_aggregate"],
            "fonti_ok":           scrape["fonti_ok"],
            "date":               date_str,
        }

        all_results.append({
            "nome":            player["nome"],
            "mercato":         player["mercato"],
            "ruolo":           player["ruolo"],
            "complessita":     player["complessita_procurement"],
            "fatturato_bn":    player["fatturato_bn"],
            "servizi_it":      player["servizi_it"],
            "fonti_list":      player["fonti"],
            "fonti_principali": player["fonti"][0] if player["fonti"] else "",
            "keywords_aggregate": scrape["keywords_aggregate"],
            "fonti_ok":        scrape["fonti_ok"],
            "fonti_ko":        scrape["fonti_ko"],
            "last_update":     scrape["last_update"],
            "delta":           delta,
        })

    save_history(history)

    xlsx_path  = build_excel(all_results, date_str)
    html_body  = build_email_html(all_results, date_str)
    send_email(xlsx_path, html_body, date_str)

    log.info("=== Fine esecuzione ===")
    return xlsx_path


# ─────────────────────────────────────────────
# SCHEDULER (opzionale – gira a ciclo continuo)
# ─────────────────────────────────────────────
if __name__ == "__main__":
    import sys

    # Esecuzione immediata singola:
    if "--once" in sys.argv:
        run()
    else:
        # Schedule alle 07:30 ogni giorno
        try:
            import schedule
            log.info("Scheduler attivo. Prossima esecuzione alle 07:30.")
            schedule.every().day.at("07:30").do(run)
            # Prima esecuzione immediata
            run()
            while True:
                schedule.run_pending()
                time.sleep(60)
        except ImportError:
            log.warning("'schedule' non installato – esecuzione singola.")
            run()
